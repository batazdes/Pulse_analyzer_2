import streamlit as st
import numpy as np
import pandas as pd
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Pulse Analyzer", page_icon="⚡", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;700&family=Syne:wght@400;700;800&display=swap');
html, body, [class*="css"] { font-family: 'Syne', sans-serif; }
.stApp { background: #0d0f14; color: #e8eaf0; }
h1 { font-family: 'Syne', sans-serif; font-weight: 800; font-size: 2.2rem; letter-spacing: -0.03em; color: #ffffff; }
h2, h3 { font-family: 'Syne', sans-serif; font-weight: 700; color: #c8cbdc; }
section[data-testid="stSidebar"] { background: #13151d !important; border-right: 1px solid #1e2130; }
section[data-testid="stSidebar"] * { color: #c8cbdc !important; }
[data-testid="metric-container"] { background: #13151d; border: 1px solid #1e2130; border-radius: 8px; padding: 1rem; }
[data-testid="metric-container"] label { color: #6b7280 !important; font-size: 0.75rem; letter-spacing: 0.08em; text-transform: uppercase; }
[data-testid="metric-container"] [data-testid="stMetricValue"] { color: #38bdf8 !important; font-family: 'JetBrains Mono', monospace; font-size: 1.4rem; }
.stDownloadButton > button { background: #1a56db !important; color: #fff !important; border: none !important; border-radius: 6px !important; font-family: 'Syne', sans-serif !important; font-weight: 700 !important; padding: 0.5rem 1.5rem !important; }
.stDownloadButton > button:hover { background: #1e40af !important; }
[data-testid="stFileUploader"] { background: #13151d; border: 1.5px dashed #2a2f45; border-radius: 10px; padding: 1.5rem; }
hr { border-color: #1e2130; }
.stSlider label { font-size: 0.8rem; color: #6b7280; text-transform: uppercase; letter-spacing: 0.06em; }
.accent-bar { height: 3px; background: linear-gradient(90deg, #1a56db, #38bdf8, #a78bfa); border-radius: 2px; margin-bottom: 2rem; }
.group-header { background: #1a1f2e; border-left: 3px solid #38bdf8; padding: 4px 12px; margin: 8px 0 2px 0; border-radius: 0 4px 4px 0; font-size: 0.8rem; color: #38bdf8; letter-spacing: 0.06em; text-transform: uppercase; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════
# Логика
# ══════════════════════════════════════════════════════════

def detect_sep(content):
    try:
        return csv.Sniffer().sniff(content[:4096], delimiters=",;\t |").delimiter
    except csv.Error:
        return ","

def is_numeric(s):
    try:
        float(s.replace(",", "."))
        return True
    except ValueError:
        return False

def load_csv_bytes(raw, col=None):
    content = raw.decode("utf-8-sig", errors="replace")
    sep = detect_sep(content)
    rows = [r for r in csv.reader(io.StringIO(content), delimiter=sep) if any(c.strip() for c in r)]
    if not rows:
        raise ValueError("CSV файл пуст")
    has_header = not all(is_numeric(c.strip()) for c in rows[0] if c.strip())
    data_rows = rows[1:] if has_header else rows
    n_cols = len(data_rows[0]) if data_rows else 0
    val_col = (1 if n_cols >= 2 else 0) if col is None else col
    time_col = 0 if n_cols >= 2 and val_col != 0 else None
    times, values = [], []
    for i, row in enumerate(data_rows):
        if val_col >= len(row):
            continue
        raw_val = row[val_col].strip().replace(",", ".")
        if not raw_val:
            continue
        try:
            values.append(float(raw_val))
        except ValueError:
            continue
        if time_col is not None and time_col < len(row):
            try:
                times.append(float(row[time_col].strip().replace(",", ".")))
            except ValueError:
                times.append(float(i))
        else:
            times.append(float(i))
    if not values:
        raise ValueError("Не удалось прочитать числовые значения")
    return np.array(times), np.array(values), sep, n_cols, has_header

def binarize(signal, thr_pct):
    lo, hi = signal.min(), signal.max()
    if hi - lo < 1e-12:
        raise ValueError("Сигнал константный")
    return signal > lo + (hi - lo) * thr_pct / 100.0

def find_pulses(binary, min_width):
    pulses, in_pulse, start = [], False, 0
    for i, val in enumerate(binary):
        if val and not in_pulse:
            in_pulse, start = True, i
        elif not val and in_pulse:
            in_pulse = False
            if i - start >= min_width:
                pulses.append((start, i - 1))
    if in_pulse and len(binary) - start >= min_width:
        pulses.append((start, len(binary) - 1))
    return pulses

def pulse_stats(signal, start, end, trim):
    s, e = start + trim, end - trim + 1
    if e - s < 2:
        return None
    seg = signal[s:e]
    return {
        "start": start, "end": end,
        "n_points": len(seg),
        "mean": float(np.mean(seg)),
        "std":  float(np.std(seg, ddof=1)),
        "raw":  seg.tolist(),
    }

def group_stats(chunk):
    means = [s["mean"] for s in chunk]
    return {
        "n_pulses": len(chunk),
        "mean": float(np.mean(means)),
        "std":  float(np.std(means, ddof=1)) if len(means) > 1 else 0.0,
    }


# ══════════════════════════════════════════════════════════
# Excel
# ══════════════════════════════════════════════════════════

H_FILL  = PatternFill("solid", start_color="1F4E79")
H_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
GR_FILL = PatternFill("solid", start_color="D6E4F0")
WH_FILL = PatternFill("solid", start_color="FFFFFF")
SP_FILL = PatternFill("solid", start_color="F0F0F0")
BL_FONT = Font(name="Arial", size=10)
CENTER  = Alignment(horizontal="center", vertical="center")
THIN    = Side(style="thin", color="BBBBBB")
BRD     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FMT5    = "0.00000"

def _c(ws, r, c, val=None, font=None, fill=None, num_fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.alignment = CENTER
    cell.border = BRD
    cell.font = font or BL_FONT
    if fill:   cell.fill = fill
    if num_fmt and isinstance(val, (int, float)):
        cell.number_format = num_fmt
    return cell

def build_excel(pulse_results, groups, group_size, save_raw):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Pulse Details"
    for ci, (h, w) in enumerate(zip(
        ["Pulse #", "Group #", "N points", "Mean", "Std Dev"],
        [12, 12, 12, 16, 16]
    ), 1):
        _c(ws1, 1, ci, h, font=H_FONT, fill=H_FILL)
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[1].height = 22

    excel_row = 2
    for i, s in enumerate(pulse_results):
        g = i // group_size + 1
        if i > 0 and i % group_size == 0:
            for ci in range(1, 6):
                cell = ws1.cell(row=excel_row, column=ci)
                cell.fill = SP_FILL
                cell.border = BRD
            ws1.row_dimensions[excel_row].height = 8
            excel_row += 1
        fill = GR_FILL if g % 2 == 0 else WH_FILL
        for ci, (v, f) in enumerate(zip(
            [i+1, g, s["n_points"] if s else "—", s["mean"] if s else "—", s["std"] if s else "—"],
            [None, None, None, FMT5, FMT5]
        ), 1):
            _c(ws1, excel_row, ci, v, fill=fill, num_fmt=f)
        excel_row += 1
    ws1.freeze_panes = "A2"

    ws2 = wb.create_sheet("Group Summary")
    for ci, (h, w) in enumerate(zip(
        ["Group #", "N pulses", "Pulse indices", "Average", "Std Dev"],
        [12, 12, 28, 16, 16]
    ), 1):
        _c(ws2, 1, ci, h, font=H_FONT, fill=H_FILL)
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[1].height = 22
    for gi, gs in enumerate(groups):
        p0   = gi * group_size + 1
        p1   = min(p0 + group_size - 1, len(pulse_results))
        fill = GR_FILL if (gi + 1) % 2 == 0 else WH_FILL
        for ci, (v, f) in enumerate(zip(
            [gi+1, gs["n_pulses"],
             ", ".join(str(x) for x in range(p0, p1+1)),
             gs["mean"], gs["std"]],
            [None, None, None, FMT5, FMT5]
        ), 1):
            _c(ws2, gi+2, ci, v, fill=fill, num_fmt=f)
    ws2.freeze_panes = "A2"

    if save_raw:
        ws3 = wb.create_sheet("Raw Values")
        max_pts = max((s["n_points"] for s in pulse_results if s), default=0)
        _c(ws3, 1, 1, "Pulse #", font=H_FONT, fill=H_FILL)
        ws3.column_dimensions["A"].width = 10
        for ci in range(1, max_pts + 1):
            _c(ws3, 1, ci+1, f"pt {ci}", font=H_FONT, fill=H_FILL)
            ws3.column_dimensions[get_column_letter(ci+1)].width = 9
        for i, s in enumerate(pulse_results):
            fill = GR_FILL if (i // group_size + 1) % 2 == 0 else WH_FILL
            _c(ws3, i+2, 1, i+1, fill=fill)
            if s:
                for ci, v in enumerate(s["raw"], 2):
                    _c(ws3, i+2, ci, round(v, 6), fill=fill, num_fmt=FMT5)
        ws3.freeze_panes = "B2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════

st.markdown('<div class="accent-bar"></div>', unsafe_allow_html=True)
st.title("⚡ Pulse Analyzer")
st.markdown("Загрузи CSV — получи статистику по импульсам и Excel-отчёт.")

with st.sidebar:
    st.markdown("## ⚙️ Параметры")
    st.markdown("---")
    trim      = st.slider("Обрезка с каждого конца (точек)", 0, 30, 5)
    group     = st.slider("Размер группы", 1, 20, 4)
    threshold = st.slider("Порог бинаризации (%)", 5, 80, 30,
                          help="% от диапазона [min, max] сигнала. Всё выше — импульс, ниже — ноль.")
    min_width = st.slider("Мин. ширина импульса (точек)", 1, 100, 5)
    st.markdown("---")
    col_auto = st.checkbox("Авто-определение столбца", value=True)
    col_idx  = st.number_input("Индекс столбца значений (0-based)", 0, 50, 1, disabled=col_auto)
    st.markdown("---")
    save_raw = st.checkbox("Сохранить сырые значения в Excel", value=False)

uploaded = st.file_uploader("📂 Загрузи CSV-файл", type=["csv", "txt"])

if uploaded is None:
    st.info("Загрузи CSV-файл, чтобы начать анализ.")
    with st.expander("📋 Поддерживаемые форматы"):
        st.markdown("""
| Формат | Пример |
|---|---|
| Один столбец | `1.23\\n4.56\\n...` |
| Время + значение | `0.001, 1.23\\n...` |
| С заголовком | `time, voltage\\n...` |
| Разделитель `;` или `Tab` | поддерживается |
        """)
    st.stop()

# ── Загрузка ──────────────────────────────────────────────
try:
    time, signal, sep_used, n_cols, has_header = load_csv_bytes(
        uploaded.read(),
        col=None if col_auto else int(col_idx)
    )
except Exception as ex:
    st.error(f"Ошибка чтения CSV: {ex}")
    st.stop()

c1, c2, c3, c4 = st.columns(4)
c1.metric("Точек", f"{len(signal):,}")
c2.metric("Столбцов", str(n_cols))
c3.metric("Разделитель", repr(sep_used))
c4.metric("Заголовок", "да" if has_header else "нет")

st.markdown("---")

# ── Анализ ────────────────────────────────────────────────
try:
    binary = binarize(signal, threshold)
    pulses = find_pulses(binary, min_width)
except ValueError as ex:
    st.error(str(ex))
    st.stop()

if len(pulses) == 0:
    st.warning("Импульсы не найдены. Попробуй уменьшить порог бинаризации или мин. ширину.")
    st.stop()

pulse_results = [pulse_stats(signal, s, e, trim) for s, e in pulses]
skipped = sum(1 for r in pulse_results if r is None)
if skipped:
    st.warning(f"{skipped} импульс(ов) пропущено — слишком короткие после обрезки. Уменьши «Обрезку».")

groups = []
for gi in range(0, len(pulse_results), group):
    chunk = [r for r in pulse_results[gi:gi + group] if r]
    if chunk:
        groups.append(group_stats(chunk))

# ── График ────────────────────────────────────────────────
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 6), sharex=True,
                                gridspec_kw={"height_ratios": [3, 1]})
fig.patch.set_facecolor("#0d0f14")
for ax in (ax1, ax2):
    ax.set_facecolor("#13151d")
    ax.tick_params(colors="#6b7280")
    for sp in ax.spines.values():
        sp.set_color("#1e2130")

ax1.plot(time, signal, color="#38bdf8", lw=1.0)
for i, (s, e) in enumerate(pulses):
    ax1.axvspan(time[s], time[e], alpha=0.10, color="#f59e0b")
    r = pulse_results[i]
    if r:
        ts = time[r["start"] + trim] if r["start"] + trim < len(time) else time[r["start"]]
        te = time[r["end"]   - trim] if r["end"]   - trim >= 0          else time[r["end"]]
        ax1.axvspan(ts, te, alpha=0.22, color="#34d399")
    ax1.text((time[s] + time[e]) / 2, signal.max(),
             f"#{i+1}", ha="center", va="bottom",
             fontsize=7, color="#f87171", fontweight="bold", fontfamily="monospace")

ax1.legend(
    handles=[mpatches.Patch(color="#f59e0b", alpha=0.4, label="Импульс"),
             mpatches.Patch(color="#34d399", alpha=0.5, label="После обрезки")],
    loc="upper right", fontsize=8, facecolor="#13151d", labelcolor="#c8cbdc"
)
ax1.set_ylabel("Значение", color="#6b7280", fontsize=9)
ax1.set_title(f"Найдено импульсов: {len(pulses)}", color="#c8cbdc", fontsize=10, loc="left")
ax1.grid(True, alpha=0.15, color="#2a2f45")

ax2.fill_between(time, binary.astype(int), step="mid", color="#a78bfa", alpha=0.6)
ax2.set_ylabel("0 / 1", color="#6b7280", fontsize=9)
ax2.set_xlabel("Время / индекс", color="#6b7280", fontsize=9)
ax2.grid(True, alpha=0.15, color="#2a2f45")

plt.tight_layout()
st.pyplot(fig)
plt.close()

st.markdown("---")

# ── Таблица импульсов по группам ──────────────────────────
st.markdown("### 📊 Детали импульсов")

for gi in range(len(groups)):
    p_start = gi * group
    p_end   = min(p_start + group, len(pulse_results))
    chunk   = pulse_results[p_start:p_end]

    st.markdown(f'<div class="group-header">Группа {gi + 1}</div>', unsafe_allow_html=True)

    rows = []
    for j, r in enumerate(chunk):
        pulse_num = p_start + j + 1
        if r:
            rows.append({
                "Pulse #":   pulse_num,
                "N points":  r["n_points"],
                "Mean":      round(r["mean"], 5),
                "Std Dev":   round(r["std"],  5),
            })
        else:
            rows.append({
                "Pulse #":  pulse_num,
                "N points": "—",
                "Mean":     "—",
                "Std Dev":  "—",
            })

    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

st.markdown("---")

# ── Сводка по группам ─────────────────────────────────────
st.markdown("### 📦 Сводка по группам")

summary_rows = []
for gi, gs in enumerate(groups):
    p0 = gi * group + 1
    p1 = min(p0 + group - 1, len(pulse_results))
    summary_rows.append({
        "Group #":       gi + 1,
        "N pulses":      gs["n_pulses"],
        "Pulse indices": ", ".join(str(x) for x in range(p0, p1 + 1)),
        "Average":       round(gs["mean"], 5),
        "Std Dev":       round(gs["std"],  5),
    })

st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

st.markdown("---")

# ── Итоговые метрики ──────────────────────────────────────
all_means = [r["mean"] for r in pulse_results if r]
m1, m2, m3 = st.columns(3)
m1.metric("Импульсов найдено", len(pulses))
m2.metric("Глобальный Average", f"{np.mean(all_means):.5f}")
m3.metric("Глобальный Std Dev", f"{np.std(all_means, ddof=1):.5f}" if len(all_means) > 1 else "—")

st.markdown("---")

# ── Скачать Excel ─────────────────────────────────────────
xlsx = build_excel(pulse_results, groups, group, save_raw)
st.download_button(
    label="⬇️  Скачать Excel (.xlsx)",
    data=xlsx,
    file_name="pulse_results.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
